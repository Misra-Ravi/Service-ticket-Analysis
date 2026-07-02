"""
JCI Service Ticket Weekly Processor
=====================================
Processes a raw ServiceNow SAPUI5 export and produces a business-ready
Excel workbook with aging, priority, and carry-forward update columns.

Usage:
    python process_tickets.py --raw <path_to_raw.xlsx> [--prev <path_to_prev_updated.xlsx>] [--out <output_path.xlsx>] [--date YYYY-MM-DD]

Examples:
    python process_tickets.py --raw "caseList 27.xlsx"
    python process_tickets.py --raw "caseList 27.xlsx" --prev "caseList 26 Updated.xlsx"
    python process_tickets.py --raw "caseList 27.xlsx" --prev "caseList 26 Updated.xlsx" --out "Weekly_2026-07-09.xlsx"
    python process_tickets.py --raw "caseList 27.xlsx" --date 2026-07-09
"""

import argparse
import sys
import os
from datetime import datetime, timezone
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Run:  pip install openpyxl")
    sys.exit(1)


# ── Constants ──────────────────────────────────────────────────────────────────

IN_SCOPE_STATUSES = {"Sent to SAP", "Customer Action", "SAP Proposed Solution"}

SAP_TO_COLOR = {
    "Very High": "Red",
    "High":      "Yellow",
    "Medium":    "Green",
    "Low":       None,
}

PRIORITY_SORT = {"VERY HIGH": 0, "HIGH": 1, "MEDIUM HIGH": 2}

NEW_COLS = ["Aging", "Last Update", "Customer Priority", "Customer Update", "SAP Update"]

# Expected column headers in the raw ServiceNow export
EXPECTED_RAW_HEADERS = [
    "CASE", "SUBJECT", "STATUS", "PRIORITY", "INSTALLATION",
    "SYSTEM NUMBER", "SYSTEM", "COMPONENT", "REPORTER ID", "REPORTER",
    "CREATOR ID", "CREATOR", "CUSTOMER ID", "CUSTOMER",
    "CREATED ON (UTC)", "UPDATED ON (UTC)", "AUTO-CONFIRM DATE",
    "SUBMITTED ON", "COMPLETED ON",
]

# Previous-week file column name variations (JCI Incidents format vs caseList Updated format)
PREV_KEY_CANDIDATES   = ["CASE", "OSS Message"]
PREV_SAP_CANDIDATES   = ["SAP Update", "MAX UPDATE"]
PREV_CUST_CANDIDATES  = ["Customer Update", "JCI UPDATE"]
PREV_SHEET_CANDIDATES = ["Current Week", "SAPUI5 Export"]


# ── Customer Priority Logic ────────────────────────────────────────────────────

def customer_priority(aging, status, sap_priority):
    """
    Derive Customer Priority using ALL THREE conditions simultaneously:
      - VERY HIGH  : Aging > 7  days  AND status in-scope AND SAP Priority = Very High (Red)
      - HIGH       : Aging > 30 days  AND status in-scope AND SAP Priority = High (Yellow)
      - MEDIUM HIGH: Aging > 60 days  AND status in-scope AND SAP Priority = Medium (Green)
    Returns None if any condition is not met.
    """
    if aging is None or status not in IN_SCOPE_STATUSES:
        return None
    color = SAP_TO_COLOR.get(str(sap_priority).strip())
    if color == "Red"    and aging > 7:   return "VERY HIGH"
    if color == "Yellow" and aging > 30:  return "HIGH"
    if color == "Green"  and aging > 60:  return "MEDIUM HIGH"
    return None


# ── File Loaders ───────────────────────────────────────────────────────────────

def load_raw_file(path):
    """Load the raw ServiceNow export. Returns (headers tuple, list of row tuples)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError(f"Raw file is empty: {path}")
    return rows[0], rows[1:]


def find_sheet(wb, candidates):
    """Return the first sheet whose name matches one of the candidates."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def find_col(headers, candidates):
    """Return the 0-based index of the first matching header candidate, or None."""
    for candidate in candidates:
        for i, h in enumerate(headers):
            if h and str(h).strip() == candidate:
                return i
    return None


def load_prev_file(path):
    """
    Load the previous week's updated file.
    Returns a dict: { case_number: {"customer_update": ..., "sap_update": ...} }
    Handles both 'caseList XX Updated.xlsx' and 'JCI Incidents' formats automatically.
    """
    if not path or not os.path.exists(path):
        return {}, None, None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_sheet(wb, PREV_SHEET_CANDIDATES)
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {}, None, None

    headers = rows[0]
    key_idx  = find_col(headers, PREV_KEY_CANDIDATES)
    sap_idx  = find_col(headers, PREV_SAP_CANDIDATES)
    cust_idx = find_col(headers, PREV_CUST_CANDIDATES)

    if key_idx is None:
        return {}, None, None

    key_name  = headers[key_idx]
    sap_name  = headers[sap_idx]  if sap_idx  is not None else None
    cust_name = headers[cust_idx] if cust_idx is not None else None

    lookup = {}
    for row in rows[1:]:
        key = str(row[key_idx]).strip() if row[key_idx] else None
        if not key:
            continue
        lookup[key] = {
            "sap_update":      row[sap_idx]  if sap_idx  is not None else None,
            "customer_update": row[cust_idx] if cust_idx is not None else None,
        }

    return lookup, key_name, (sap_name, cust_name)


# ── Row Processor ──────────────────────────────────────────────────────────────

def process_rows(raw_hdr, raw_data, prev_lookup, today):
    """
    Process all raw rows. Returns list of processed rows (original cols + 5 new cols).
    Each processed row is a list: [...original_data..., aging, last_update, customer_priority,
                                   customer_update, sap_update]
    """
    def idx(name):
        return raw_hdr.index(name)

    COL_CASE     = idx("CASE")
    COL_STATUS   = idx("STATUS")
    COL_PRIORITY = idx("PRIORITY")
    COL_CREATED  = idx("CREATED ON (UTC)")
    COL_UPDATED  = idx("UPDATED ON (UTC)")

    issues = []
    processed = []

    for row in raw_data:
        case   = str(row[COL_CASE]).strip()   if row[COL_CASE]     else ""
        status = str(row[COL_STATUS]).strip()  if row[COL_STATUS]   else ""
        prio   = str(row[COL_PRIORITY]).strip() if row[COL_PRIORITY] else ""

        # Aging
        cre = row[COL_CREATED]
        if isinstance(cre, datetime):
            cre   = cre.replace(tzinfo=timezone.utc) if cre.tzinfo is None else cre
            aging = (today - cre).days
        else:
            aging = None
            issues.append(f"CASE {case}: CREATED ON (UTC) blank/invalid — Aging left blank")

        # Last Update
        upd = row[COL_UPDATED]
        if isinstance(upd, datetime):
            upd      = upd.replace(tzinfo=timezone.utc) if upd.tzinfo is None else upd
            last_upd = (today - upd).days
        else:
            last_upd = None
            issues.append(f"CASE {case}: UPDATED ON (UTC) blank/invalid — Last Update left blank")

        cp   = customer_priority(aging, status, prio)
        prev = prev_lookup.get(case, {})

        processed.append(
            list(row) + [
                aging,
                last_upd,
                cp,
                prev.get("customer_update"),
                prev.get("sap_update"),
            ]
        )

    return processed, issues


# ── Excel Writer ───────────────────────────────────────────────────────────────

def write_output(final_hdr, all_rows, in_scope_count, out_path):
    """Write the final formatted Excel workbook."""

    N = len(final_hdr) - len(NEW_COLS)   # number of original columns

    THIN      = Side(style="thin", color="AAAAAA")
    BORDER    = Border(left=THIN, right=THIN, bottom=THIN, top=THIN)
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    NEW_FILL  = PatternFill("solid", fgColor="FFD966")
    NEW_FONT  = Font(bold=True, color="000000", size=11)
    EVEN_FILL = PatternFill("solid", fgColor="EAF1FB")
    ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
    LOCKED    = Protection(locked=True)
    UNLOCKED  = Protection(locked=False)

    CU_COL  = final_hdr.index("Customer Update") + 1   # 1-based
    SAP_COL = final_hdr.index("SAP Update") + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Current Week"

    # Header row
    ws.append(final_hdr)
    for ci, cell in enumerate(ws[1], start=1):
        is_new     = final_hdr[ci-1] in NEW_COLS
        cell.fill  = NEW_FILL if is_new else HDR_FILL
        cell.font  = NEW_FONT if is_new else HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        cell.protection = LOCKED
    ws.row_dimensions[1].height = 32

    # Data rows
    for ri, row in enumerate(all_rows, start=1):
        ws.append(row)
        excel_row = ri + 1
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, cell in enumerate(ws[excel_row], start=1):
            cell.fill   = fill
            cell.border = BORDER
            if ci > N:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.protection = UNLOCKED if ci in (CU_COL, SAP_COL) else LOCKED

    # Freeze columns A+B and header row (C2)
    ws.freeze_panes = "C2"

    # Auto-filter on header row
    last_col_letter = get_column_letter(len(final_hdr))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Sheet protection — only Customer Update and SAP Update editable
    ws.protection.sheet              = True
    ws.protection.password           = "JCI2026"
    ws.protection.selectLockedCells  = False
    ws.protection.selectUnlockedCells = False
    ws.protection.autoFilter         = False
    ws.protection.sort               = False

    # Column widths
    widths = {
        "CASE": 18, "SUBJECT": 50, "STATUS": 22, "PRIORITY": 12,
        "INSTALLATION": 35, "SYSTEM NUMBER": 15, "SYSTEM": 20,
        "COMPONENT": 40, "REPORTER ID": 14, "REPORTER": 22,
        "CREATOR ID": 14, "CREATOR": 22, "CUSTOMER ID": 14, "CUSTOMER": 28,
        "CREATED ON (UTC)": 20, "UPDATED ON (UTC)": 20, "AUTO-CONFIRM DATE": 18,
        "SUBMITTED ON": 15, "COMPLETED ON": 15,
        "Aging": 10, "Last Update": 13, "Customer Priority": 18,
        "Customer Update": 35, "SAP Update": 35,
    }
    for ci, h in enumerate(final_hdr, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)

    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="JCI Service Ticket Weekly Processor",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--raw",  required=True, help="Path to current week raw ServiceNow export (.xlsx)")
    parser.add_argument("--prev", default=None,  help="Path to previous week updated file (.xlsx) for carry-forward")
    parser.add_argument("--out",  default=None,  help="Output file path (default: <raw_name> Updated.xlsx)")
    parser.add_argument("--date", default=None,  help="Override today's date as YYYY-MM-DD (default: system date)")
    args = parser.parse_args()

    # ── Validate inputs ──
    if not os.path.exists(args.raw):
        print(f"ERROR: Raw file not found: {args.raw}")
        sys.exit(1)

    # ── Determine today ──
    if args.date:
        try:
            today = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            print(f"ERROR: Invalid date format '{args.date}'. Use YYYY-MM-DD.")
            sys.exit(1)
    else:
        today = datetime.now(tz=timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

    # ── Determine output path ──
    if args.out:
        out_path = args.out
    else:
        base = os.path.splitext(os.path.basename(args.raw))[0]
        out_dir = os.path.dirname(os.path.abspath(args.raw))
        out_path = os.path.join(out_dir, f"{base} Updated.xlsx")

    print(f"\n{'='*60}")
    print(f"  JCI Service Ticket Weekly Processor")
    print(f"{'='*60}")
    print(f"  Raw file   : {args.raw}")
    print(f"  Prev file  : {args.prev or '(none)'}")
    print(f"  Output     : {out_path}")
    print(f"  Today      : {today.strftime('%Y-%m-%d')}")
    print(f"{'='*60}\n")

    # ── Load raw file ──
    print("Loading raw file...")
    raw_hdr, raw_data = load_raw_file(args.raw)
    print(f"  {len(raw_data)} data rows | {len(raw_hdr)} columns")

    # ── Validate raw headers ──
    missing = [h for h in EXPECTED_RAW_HEADERS if h not in raw_hdr]
    if missing:
        print(f"  WARNING: Expected columns not found in raw file: {missing}")

    # ── Load previous file ──
    print("Loading previous week file...")
    prev_lookup, key_col_used, update_cols_used = load_prev_file(args.prev)
    if prev_lookup:
        print(f"  {len(prev_lookup)} records in carry-forward lookup")
        print(f"  Key column used   : {key_col_used}")
        print(f"  Update cols found : SAP='{update_cols_used[0]}', Customer='{update_cols_used[1]}'")
    else:
        print("  No previous file — Customer Update and SAP Update will be blank")

    # ── Process rows ──
    print("Processing rows...")
    processed, issues = process_rows(raw_hdr, raw_data, prev_lookup, today)

    N = len(raw_hdr)
    AGING_I  = N
    CP_I     = N + 2

    in_scope  = [r for r in processed if r[2] in IN_SCOPE_STATUSES]   # col 2 = STATUS
    out_scope = [r for r in processed if r[2] not in IN_SCOPE_STATUSES]

    in_scope.sort(key=lambda r: (-(r[AGING_I] or -1), PRIORITY_SORT.get(r[CP_I], 99)))
    all_rows = in_scope + out_scope

    # ── Build final headers ──
    final_hdr = list(raw_hdr) + NEW_COLS

    # ── Write output ──
    print("Writing output file...")
    write_output(final_hdr, all_rows, len(in_scope), out_path)

    # ── Summary ──
    matched   = sum(1 for r in in_scope if prev_lookup.get(str(r[0]).strip()))
    cp_counts = dict(Counter(r[CP_I] or "blank" for r in in_scope))
    status_dist = dict(Counter(r[2] for r in processed))

    print(f"\n{'='*60}")
    print(f"  PROCESSING SUMMARY")
    print(f"{'='*60}")
    print(f"  Total raw rows           : {len(raw_data)}")
    print(f"  In-scope (visible, top)  : {len(in_scope)}")
    print(f"  Out-of-scope (below)     : {len(out_scope)}")
    print(f"  Matched carry-forward    : {matched}")
    print(f"  New rows (blank logs)    : {len(in_scope) - matched}")
    print(f"  Date issues              : {len(issues)}")
    if issues:
        for issue in issues:
            print(f"    • {issue}")
    print(f"\n  Status distribution:")
    for status, count in sorted(status_dist.items(), key=lambda x: -x[1]):
        marker = " ← IN SCOPE" if status in IN_SCOPE_STATUSES else ""
        print(f"    {status:<35} {count:>4}{marker}")
    print(f"\n  Customer Priority (in-scope rows):")
    for label in ["VERY HIGH", "HIGH", "MEDIUM HIGH", "blank"]:
        print(f"    {label:<15} : {cp_counts.get(label, 0)}")
    print(f"\n  Color/category source    : SAP PRIORITY mapped to color tiers")
    print(f"    Very High → Red  → VERY HIGH   (Aging > 7 days)")
    print(f"    High      → Yellow → HIGH      (Aging > 30 days)")
    print(f"    Medium    → Green → MEDIUM HIGH (Aging > 60 days)")
    print(f"    Low       → No priority assigned")
    print(f"\n  Output saved to: {out_path}")
    print(f"{'='*60}\n")

    if issues:
        print(f"  ⚠  {len(issues)} date issue(s) found — see above")

    return out_path


if __name__ == "__main__":
    main()
